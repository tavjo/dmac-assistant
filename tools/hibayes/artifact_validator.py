"""T1.1 — Stage A: Artifact Validity validator.

Reads `evidence/headless/<run>/manifest.json` (per locked DD-48 as plain dict —
do NOT call `RawRunManifest.model_validate`), rebases per-summary artifact paths
against a caller-supplied `--artifact-root` per locked DD-29, runs per-task-family
structural validators (GEO `.xlsx` per DD-19+DD-38, nf-core RNA-seq CSV per DD-17,
nf-core scRNA-seq CSV per DD-18, SVG per DD-20), and emits
`hibayes_artifact_validity.csv` (29 columns per locked §5.1).

DD-25 caveat: running against a manifest emitted without `--scratch-dir`
(tempdir-mode) will produce a CSV in which every expected-artifact row is
`Missing`, indistinguishable from real failure. See DD25_TEMPDIR_WARNING constant.

plan-DD-03 guard: NotImplementedError raised on any multi-file artifact summary.
DL-014 hardening: CLI --help text includes DD25_TEMPDIR_WARNING verbatim.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from tools.hibayes.enums import ArtifactKind, ArtifactStatus
from tools.hibayes.exporter import parse_query_id


__all__ = [
    "CSV_HEADER_29",
    "DD25_TEMPDIR_WARNING",
    "ValidatorResult",
    "_load_and_validate_geo_template",  # Pass 4 D1: exposed for direct pinning tests
    "classify_artifact_kind",
    "main",
    "rebase_artifact_path",
    "run_stage_a",
    "validate_geo_xlsx",
    "validate_nfcore_rnaseq_csv",
    "validate_nfcore_scrnaseq_csv",
    "validate_svg",
]


# Locked §5.1 — 29 columns in exact order.
CSV_HEADER_29: tuple[str, ...] = (
    "run_id",
    "query_id",
    "task_family",
    "artifact_eval_id",
    "artifact_expected",
    "expected_artifact_kind",
    "artifact_declared",
    "artifact_path",
    "artifact_basename",
    "artifact_ext",
    "runtime_success",
    "failure_mode",
    "artifact_exists",
    "artifact_accessible",
    "file_size_bytes",
    "parser_used",
    "parse_success",
    "sheet_count",
    "row_count",
    "column_count",
    "nonempty_cell_count",
    "null_cell_fraction",
    "required_fields_present",
    "required_fields_complete",
    "missing_required_fields",
    "all_required_rows_complete",
    "artifact_validity_status",
    "artifact_success",
    "validation_notes",
)


# DL-014: verbatim warning string in --help and module docstring.
DD25_TEMPDIR_WARNING = (
    "WARNING (locked DD-25): Running Stage A against a manifest emitted without "
    "`--scratch-dir` (tempdir-mode) will produce a `hibayes_artifact_validity.csv` "
    "in which every expected-artifact row is `Missing`, indistinguishable from real "
    "failure. Stage A cannot detect tempdir-mode-vs-fixed-scratch-mode from the "
    "manifest itself. Verify the upstream run used `--scratch-dir` before drawing "
    "inferences from a 100%-Missing Stage A CSV."
)


# nf-core RNA-seq enum per locked DD-17.
_NFCORE_RNASEQ_STRANDEDNESS_ENUM = frozenset({"unstranded", "forward", "reverse", "auto"})
_NFCORE_RNASEQ_REQUIRED = ("sample", "fastq_1", "fastq_2", "strandedness")
_NFCORE_SCRNASEQ_REQUIRED = ("sample", "fastq_1", "fastq_2")


@dataclass
class ValidatorResult:
    """Per-row validator output. All fields map 1:1 to CSV columns."""

    status: ArtifactStatus
    parser_used: str | None = None
    parse_success: bool | None = None
    sheet_count: int | None = None
    row_count: int | None = None
    column_count: int | None = None
    nonempty_cell_count: int | None = None
    null_cell_fraction: float | None = None
    required_fields_present: bool | None = None
    required_fields_complete: bool | None = None
    missing_required_fields: str | None = None
    all_required_rows_complete: bool | None = None
    validation_notes: str = ""


def rebase_artifact_path(
    *,
    manifest_artifact_abs_path: str,
    artifact_root: Path,
    qid: str,
) -> Path:
    """Locked DD-29 step 2: rebased open-path = <artifact-root>/<qid>/<basename>."""
    basename = Path(manifest_artifact_abs_path).name
    return artifact_root / qid / basename


def classify_artifact_kind(task_family: str, basename: str | None) -> ArtifactKind:
    """Locked DD-34: per-task-family → ArtifactKind routing.

    Single-file routing for v1 per plan-DD-03 (multi-file deferred).

    D2 fix (MED): `Report-NFCORE` discriminates bulk vs single-cell via basename.
    If the basename contains "scrnaseq" or "scrna" (case-insensitive), route to
    `NFCORE_SCRNASEQ_CSV` (locked DD-18). Otherwise fall back to bulk
    `NFCORE_RNASEQ_CSV` (locked DD-17). This makes
    `validate_nfcore_scrnaseq_csv` reachable from the dispatch.
    """
    if task_family == "Report-GEO":
        return ArtifactKind.GEO_XLSX
    if task_family == "Report-NFCORE":
        if basename:
            lowered = basename.lower()
            if "scrnaseq" in lowered or "scrna" in lowered:
                return ArtifactKind.NFCORE_SCRNASEQ_CSV
        return ArtifactKind.NFCORE_RNASEQ_CSV  # default bulk RNA-seq
    if task_family == "Report-SRA":
        return ArtifactKind.SRA_PACKAGE
    if task_family == "Report-PRIDE":
        return ArtifactKind.PRIDE_PACKAGE
    if basename and basename.lower().endswith(".svg"):
        return ArtifactKind.SVG_CHART
    return ArtifactKind.NONE_EXPECTED


def validate_geo_xlsx(
    *,
    file_path: Path,
    geo_template_path: Path | None,
) -> ValidatorResult:
    """Locked DD-19 + DD-38: GEO `.xlsx` validation.

    Open via openpyxl; report structural counts; enforce `*`-prefixed required
    fields scoped to `samples[0]` in the GEO template if provided.
    """
    if not file_path.exists():
        return ValidatorResult(
            status=ArtifactStatus.Missing,
            validation_notes=f"file not at {file_path}",
        )
    try:
        import openpyxl
    except ImportError:
        return ValidatorResult(
            status=ArtifactStatus.Unreadable,
            parser_used="openpyxl",
            parse_success=False,
            validation_notes="openpyxl not importable",
        )

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        return ValidatorResult(
            status=ArtifactStatus.Unreadable,
            parser_used="openpyxl",
            parse_success=False,
            validation_notes=f"openpyxl error: {type(exc).__name__}",
        )

    sheet_count = len(wb.sheetnames)
    if sheet_count == 0:
        wb.close()
        return ValidatorResult(
            status=ArtifactStatus.SchemaInvalid,
            parser_used="openpyxl",
            parse_success=True,
            sheet_count=0,
            validation_notes="workbook has zero sheets",
        )

    ws = wb[wb.sheetnames[0]]
    row_count = ws.max_row or 0
    column_count = ws.max_column or 0
    nonempty_cells = 0
    total_cells = 0
    # Pass 3 D1 fix (HIGH): in read_only=True mode the underlying archive is
    # closed by wb.close(), and any subsequent iter_rows on `ws` raises because
    # the ReadOnlyWorksheet streams from the still-open archive. Collect ALL
    # cell text we need (structural counts, the header row, AND per-row data
    # for the DD-38 step-4 per-row non-null check) in a SINGLE pass BEFORE
    # wb.close().
    #
    # Pass 4 D2 fix (HIGH): the prior version flattened every cell into a
    # single `workbook_text` string and did a substring scan for required
    # field names. Locked DD-38 step 4 instead mandates two independent
    # checks: (i) each single-`*`-prefixed required field is present as a
    # **column header** in the `samples`-equivalent sheet, AND (ii) each
    # required column is **non-null on every data row**. Substring scan
    # conflates header presence with arbitrary mentions (e.g., `*title` in a
    # `study`-block free-form cell falsely satisfies "present"), and
    # collapsing `required_complete = required_present` discards the per-row
    # null check entirely. Below we capture row 1 as the header row and the
    # remaining rows as data rows for the per-row null check.
    header_row: list[str] = []
    data_rows: list[tuple[Any, ...]] = []
    row_index = 0
    for row in ws.iter_rows(values_only=True):
        if row_index == 0:
            header_row = ["" if v is None else str(v) for v in row]
        else:
            data_rows.append(row)
        for value in row:
            total_cells += 1
            if value not in (None, ""):
                nonempty_cells += 1
        row_index += 1
    null_fraction = (
        (total_cells - nonempty_cells) / total_cells if total_cells > 0 else None
    )
    wb.close()

    # Required-field enforcement against GEO-updated.json template, samples[0] scope.
    # Pass 4 D1 fix (HIGH): the locked DD-38 schema sanity check + cardinality
    # check are performed by the module-level helper `_load_and_validate_geo_template`
    # below; that helper raises `SystemExit(2)` with the verbatim DD-38 error
    # message if `samples` is missing, `samples[0]` is malformed, the
    # single-`*`-prefixed key count is outside `[3, 25]`, or the template JSON
    # itself is malformed. By the time we reach this branch, the template is
    # either None (no enforcement) or already validated.
    required_present: bool | None = None
    required_complete: bool | None = None
    missing_fields_str: str | None = None
    all_required_rows_complete_val: bool | None = None
    if geo_template_path is not None and geo_template_path.is_file():
        # Will raise SystemExit on malformed template (DD-38 fail-loud).
        required = _load_and_validate_geo_template(geo_template_path)
        # DD-38 step 4 part (i): column-header presence check against row 1.
        # Compare case-insensitively against header cell text (stripped of
        # whitespace) so a workbook header `Library Name` matches a template
        # required field `*library name`.
        header_lower = [h.strip().lower() for h in header_row]
        missing_headers = [r for r in required if r.lower() not in header_lower]
        required_present = len(missing_headers) == 0
        # DD-38 step 4 part (ii): per-row non-null check. For each required
        # field whose header IS present, walk every data row and confirm the
        # corresponding cell is non-null. Required fields whose headers are
        # MISSING cannot be per-row-checked — they are already counted under
        # `missing_headers` and the row-completeness boolean reflects only the
        # present-required-columns dimension (consistent with DD-17/DD-18,
        # which keep `required_fields_present` and `all_required_rows_complete`
        # as two independent booleans).
        rows_complete = True
        per_row_missing: list[str] = []
        for r in required:
            if r.lower() in header_lower:
                col_idx = header_lower.index(r.lower())
                for data_row in data_rows:
                    if col_idx >= len(data_row):
                        rows_complete = False
                        per_row_missing.append(r)
                        break
                    cell = data_row[col_idx]
                    if cell is None or (isinstance(cell, str) and cell.strip() == ""):
                        rows_complete = False
                        per_row_missing.append(r)
                        break
        required_complete = required_present and rows_complete
        all_required_rows_complete_val = rows_complete if required_present else False
        # Compose `missing_required_fields`: header-absent fields first, then
        # any present-but-incomplete fields. Both contribute to the column.
        all_missing = list(missing_headers) + per_row_missing
        if all_missing:
            missing_fields_str = ";".join(all_missing)

    # Status decision: any missing header OR any incomplete row → SchemaInvalid
    # (per DD-38 + §5.1 row 24/26 semantics: header absent or data null on a
    # required column is a structural failure, not a "valid" artifact).
    if required_present is False:
        status = ArtifactStatus.SchemaInvalid
    elif required_complete is False:
        status = ArtifactStatus.Incomplete
    else:
        status = ArtifactStatus.Valid
    return ValidatorResult(
        status=status,
        parser_used="openpyxl",
        parse_success=True,
        sheet_count=sheet_count,
        row_count=row_count,
        column_count=column_count,
        nonempty_cell_count=nonempty_cells,
        null_cell_fraction=null_fraction,
        required_fields_present=required_present,
        required_fields_complete=required_complete,
        missing_required_fields=missing_fields_str,
        all_required_rows_complete=all_required_rows_complete_val,
    )


def _load_and_validate_geo_template(geo_template_path: Path) -> list[str]:
    """Pass 4 D1 fix (HIGH): fail-loud GEO template schema + cardinality check.

    Locked DD-38 step 2 + step 3 cardinality clause:
      "Stage A's GEO validator SHALL assert that the loaded structure contains
       a `samples` key whose value is a non-empty list whose first element is
       a dict (i.e., `samples[0]` exists and is a dict), AND that >=1 single-
       `*`-prefixed key is present under `samples[0]`."
      "the count of single-`*`-prefixed keys extracted from `samples[0]` MUST
       be within a reasonable range (e.g., 3-25). If the count is outside
       this range, Stage A SHALL error with the same fail-loud message ..."

    On any of the following conditions, raise `SystemExit(2)` with the verbatim
    DD-38 error message + name the offending template path on stderr:
      - template JSON malformed (`json.JSONDecodeError`)
      - top-level object is not a dict
      - `samples` key missing / not a list / empty
      - `samples[0]` not a dict
      - count of single-`*`-prefixed keys in `samples[0]` < 3 or > 25

    Returns the list of single-`*`-prefixed required field names (each with
    the leading `*` stripped, e.g., `*library name` -> `library name`).
    """
    dd38_message = (
        f"GEO template at {geo_template_path} does not match the expected "
        f"schema; was vendor/chat_nextseek/ updated? See DD-38."
    )

    def _fail(detail: str) -> None:
        sys.stderr.write(f"ERROR: {dd38_message} ({detail})\n")
        raise SystemExit(2)

    try:
        template = json.loads(geo_template_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        _fail(f"JSONDecodeError: {exc}")
        return []  # unreachable; for type-checker

    if not isinstance(template, dict):
        _fail(f"top-level structure is {type(template).__name__}, expected dict")
    samples = template.get("samples")  # type: ignore[union-attr]
    if not isinstance(samples, list):
        _fail("`samples` key missing or not a list")
    if len(samples) == 0:  # type: ignore[arg-type]
        _fail("`samples` list is empty")
    first_sample = samples[0]  # type: ignore[index]
    if not isinstance(first_sample, dict):
        _fail(
            f"`samples[0]` is {type(first_sample).__name__}, expected dict"
        )

    required: list[str] = []
    for key in first_sample.keys():  # type: ignore[union-attr]
        if (
            isinstance(key, str)
            and key.startswith("*")
            and not key.startswith("**")
        ):
            required.append(key[1:])

    # Cardinality sanity check (DD-38 step 3): 3 <= count <= 25.
    if len(required) < 3 or len(required) > 25:
        _fail(
            f"single-`*`-prefixed key count in samples[0] is {len(required)}, "
            f"expected 3-25"
        )
    return required


def validate_nfcore_rnaseq_csv(*, file_path: Path) -> ValidatorResult:
    """Locked DD-17: nf-core RNA-seq samplesheet validation."""
    if not file_path.exists():
        return ValidatorResult(status=ArtifactStatus.Missing)
    try:
        import pandas as pd

        df = pd.read_csv(file_path)
    except Exception as exc:  # noqa: BLE001
        return ValidatorResult(
            status=ArtifactStatus.Unreadable,
            parser_used="pandas_csv",
            parse_success=False,
            validation_notes=f"pandas error: {type(exc).__name__}",
        )

    missing_cols = [c for c in _NFCORE_RNASEQ_REQUIRED if c not in df.columns]
    if missing_cols:
        return ValidatorResult(
            status=ArtifactStatus.SchemaInvalid,
            parser_used="pandas_csv",
            parse_success=True,
            row_count=len(df),
            column_count=len(df.columns),
            required_fields_present=False,
            required_fields_complete=False,
            missing_required_fields=";".join(missing_cols),
        )

    # Required columns present; check completeness + strandedness enum.
    nulls_per_required = {
        c: int(df[c].isna().sum() + (df[c].astype(str).str.strip() == "").sum())
        for c in _NFCORE_RNASEQ_REQUIRED
    }
    incomplete_cols = [c for c, n in nulls_per_required.items() if n > 0]
    enum_violation = False
    if "strandedness" in df.columns:
        bad = ~df["strandedness"].astype(str).isin(_NFCORE_RNASEQ_STRANDEDNESS_ENUM)
        enum_violation = bool(bad.any())

    if enum_violation:
        return ValidatorResult(
            status=ArtifactStatus.SchemaInvalid,
            parser_used="pandas_csv",
            parse_success=True,
            row_count=len(df),
            column_count=len(df.columns),
            required_fields_present=True,
            required_fields_complete=len(incomplete_cols) == 0,
            missing_required_fields="strandedness:enum_violation"
            + (";" + ";".join(incomplete_cols) if incomplete_cols else ""),
            all_required_rows_complete=False,
        )

    if incomplete_cols:
        return ValidatorResult(
            status=ArtifactStatus.Incomplete,
            parser_used="pandas_csv",
            parse_success=True,
            row_count=len(df),
            column_count=len(df.columns),
            required_fields_present=True,
            required_fields_complete=False,
            missing_required_fields=";".join(incomplete_cols),
            all_required_rows_complete=False,
        )

    return ValidatorResult(
        status=ArtifactStatus.Valid,
        parser_used="pandas_csv",
        parse_success=True,
        row_count=len(df),
        column_count=len(df.columns),
        required_fields_present=True,
        required_fields_complete=True,
        all_required_rows_complete=True,
    )


def validate_nfcore_scrnaseq_csv(*, file_path: Path) -> ValidatorResult:
    """Locked DD-18: nf-core scRNA-seq samplesheet validation (first 3 cols required)."""
    if not file_path.exists():
        return ValidatorResult(status=ArtifactStatus.Missing)
    try:
        import pandas as pd

        df = pd.read_csv(file_path)
    except Exception as exc:  # noqa: BLE001
        return ValidatorResult(
            status=ArtifactStatus.Unreadable,
            parser_used="pandas_csv",
            parse_success=False,
            validation_notes=f"pandas error: {type(exc).__name__}",
        )

    missing_cols = [c for c in _NFCORE_SCRNASEQ_REQUIRED if c not in df.columns]
    if missing_cols:
        return ValidatorResult(
            status=ArtifactStatus.SchemaInvalid,
            parser_used="pandas_csv",
            parse_success=True,
            row_count=len(df),
            column_count=len(df.columns),
            required_fields_present=False,
            missing_required_fields=";".join(missing_cols),
        )
    incomplete_cols = [
        c for c in _NFCORE_SCRNASEQ_REQUIRED if df[c].isna().any() or (df[c].astype(str).str.strip() == "").any()
    ]
    if incomplete_cols:
        return ValidatorResult(
            status=ArtifactStatus.Incomplete,
            parser_used="pandas_csv",
            parse_success=True,
            row_count=len(df),
            column_count=len(df.columns),
            required_fields_present=True,
            required_fields_complete=False,
            missing_required_fields=";".join(incomplete_cols),
            all_required_rows_complete=False,
        )
    return ValidatorResult(
        status=ArtifactStatus.Valid,
        parser_used="pandas_csv",
        parse_success=True,
        row_count=len(df),
        column_count=len(df.columns),
        required_fields_present=True,
        required_fields_complete=True,
        all_required_rows_complete=True,
    )


def validate_svg(*, file_path: Path) -> ValidatorResult:
    """Locked DD-20: SVG validation (exists, size>0, well-formed XML, root is <svg>)."""
    if not file_path.exists():
        return ValidatorResult(status=ArtifactStatus.Missing)
    if file_path.stat().st_size == 0:
        # Pass 3 D1 fix (BLOCKER): `ValidatorResult` has no `file_size_bytes`
        # field per the dataclass definition above (it is a CSV column populated
        # in `run_stage_a`, not a `ValidatorResult` field). Passing it here
        # raised `TypeError: ValidatorResult.__init__() got an unexpected
        # keyword argument 'file_size_bytes'` on every zero-byte SVG and broke
        # the `test_validate_svg_zero_size` pinning test.
        return ValidatorResult(
            status=ArtifactStatus.Unreadable,
            parser_used="xml_etree",
            parse_success=False,
            validation_notes="empty file",
        )
    # Pass 3 D3 fix (HIGH): widen the catch surface to `(ET.ParseError, OSError)`
    # so a permission-denied SVG (`chmod 0o000`) raised by `ET.parse` — which
    # passes `Path.exists()` and `Path.stat()` (metadata-only) but raises
    # `PermissionError` (an `OSError`) on file-open — is captured as
    # `Unreadable` rather than escaping uncaught and crashing `run_stage_a`.
    # Sibling validators use broad `except Exception`; `OSError` is the
    # principled narrower scope matching DD-20's "unreadable file" semantics
    # without swallowing programmer errors.
    try:
        tree = ET.parse(file_path)
    except (ET.ParseError, OSError) as exc:
        return ValidatorResult(
            status=ArtifactStatus.Unreadable,
            parser_used="xml_etree",
            parse_success=False,
            validation_notes=f"xml parse/io error: {type(exc).__name__}: {exc}",
        )
    root = tree.getroot()
    # Accept both `<svg>` and `{namespace}svg` roots.
    tag = root.tag
    if tag.split("}", 1)[-1] != "svg":
        return ValidatorResult(
            status=ArtifactStatus.SchemaInvalid,
            parser_used="xml_etree",
            parse_success=True,
            validation_notes=f"root tag not <svg>: {tag}",
        )
    return ValidatorResult(
        status=ArtifactStatus.Valid,
        parser_used="xml_etree",
        parse_success=True,
    )


def _derive_failure_mode(summary: dict[str, Any]) -> str:
    """Mirror tools.hibayes.exporter.derive_failure_mode behavior for raw summaries."""
    if summary.get("timed_out"):
        return "timeout"
    if summary.get("is_error"):
        return "error"
    if not summary.get("answer_provided", True):
        return "no_answer"
    return "none"


def run_stage_a(
    *,
    manifest_path: Path,
    artifact_root: Path,
    geo_template_path: Path | None,
    out_csv_path: Path,
    ignore_rebase_failures: bool,
) -> int:
    """Run Stage A end-to-end. Returns process exit code.

    Locked DD-48: read manifest as plain dict.
    plan-DD-03: NotImplementedError on multi-file summaries.
    Locked DD-29 step 5: bulk-rebase-failure gate (non-zero exit unless override).
    """
    with manifest_path.open("r", encoding="utf-8") as fh:
        manifest = json.load(fh)
    run_id = manifest.get("run_id", "")
    summaries = manifest.get("summaries", [])

    # Multi-file guard.
    for s in summaries:
        if len(s.get("artifacts", [])) > 1:
            raise NotImplementedError(
                f"multi-file artifacts not yet supported (plan-DD-03). "
                f"summaries[*].artifacts has {len(s['artifacts'])} entries for "
                f"query_id={s.get('query_id', '?')}."
            )

    rows: list[list[Any]] = []
    rebase_attempted = 0
    rebase_failed = 0

    for s in summaries:
        qid = s.get("query_id", "")
        task_family, _, _ = parse_query_id(qid)
        # NOTE: `expected_behavior` is a Stage B/C concern per DD-30 — Stage A's CSV does
        # NOT carry an `expected_behavior` column (see §5.1 29-column header). Stage B
        # joins `expected_behavior_rule(task_family)` onto its own CSV; Stage A only
        # consumes `task_family`.
        # artifact_expected per DD-30: True iff task_family maps to GenerateArtifact (or Edge SVG case).
        # For Stage A's purposes, "artifact_expected" comes from `expected_artifact_kind != NONE_EXPECTED`.
        artifacts = s.get("artifacts", [])
        artifact_declared = len(artifacts) > 0

        if artifact_declared:
            assert len(artifacts) == 1  # guarded above
            manifest_path_str = artifacts[0]
            basename = Path(manifest_path_str).name
            expected_kind = classify_artifact_kind(task_family, basename)
            artifact_expected = expected_kind != ArtifactKind.NONE_EXPECTED
            rebased = rebase_artifact_path(
                manifest_artifact_abs_path=manifest_path_str,
                artifact_root=artifact_root,
                qid=qid,
            )
            rebase_attempted += 1
            if not rebased.exists():
                rebase_failed += 1
        else:
            manifest_path_str = ""
            basename = None
            # No declared artifact → infer from task_family.
            family_to_kind = {
                "Report-GEO": ArtifactKind.GEO_XLSX,
                "Report-NFCORE": ArtifactKind.NFCORE_RNASEQ_CSV,
                "Report-PRIDE": ArtifactKind.PRIDE_PACKAGE,
                "Report-SRA": ArtifactKind.SRA_PACKAGE,
            }
            expected_kind = family_to_kind.get(task_family, ArtifactKind.NONE_EXPECTED)
            artifact_expected = expected_kind != ArtifactKind.NONE_EXPECTED
            rebased = None

        runtime_success = (
            s.get("answer_provided", False)
            and not s.get("is_error", False)
            and not s.get("timed_out", False)
        )
        failure_mode = _derive_failure_mode(s)
        eval_id = (
            f"{qid}::0" if artifact_declared else f"{qid}::expected"
        )

        result: ValidatorResult
        artifact_exists: bool | None
        artifact_accessible: bool | None
        file_size: int | None
        if not artifact_expected and not artifact_declared:
            # NotExpected — task_family that doesn't produce artifacts AND none declared.
            result = ValidatorResult(status=ArtifactStatus.NotExpected)
            artifact_exists = None
            artifact_accessible = None
            file_size = None
        elif artifact_expected and not artifact_declared:
            # Expected but absent.
            # Pass-5 D1 fix: locked DD-36 — when `runtime_success == false` AND no file
            # exists, the row is `RuntimeFailed` (NOT `Missing`). When runtime succeeded
            # but the artifact is absent, the row remains `Missing` as before.
            if not runtime_success:
                result = ValidatorResult(status=ArtifactStatus.RuntimeFailed)
            else:
                result = ValidatorResult(status=ArtifactStatus.Missing)
            artifact_exists = False
            artifact_accessible = False
            file_size = None
        else:
            # artifact_declared (rebased path computed).
            assert rebased is not None
            artifact_exists = rebased.exists()
            # Locked §5.1 row 14: `os.access(..., os.R_OK)` on the rebased open-path.
            # A file that exists but is unreadable (permission-denied) MUST report
            # artifact_accessible=False so downstream Stage B/C can distinguish
            # "Missing" from "Inaccessible".
            artifact_accessible = (
                os.access(rebased, os.R_OK) if artifact_exists else False
            )
            file_size = rebased.stat().st_size if artifact_exists else None
            if not artifact_exists:
                # Pass-5 D1 fix: locked DD-36 — `runtime_success == false` AND no file on
                # disk → `RuntimeFailed`. Otherwise `Missing`.
                if not runtime_success:
                    result = ValidatorResult(
                        status=ArtifactStatus.RuntimeFailed,
                        validation_notes=f"rebased path not found: {rebased}",
                    )
                else:
                    result = ValidatorResult(
                        status=ArtifactStatus.Missing,
                        validation_notes=f"rebased path not found: {rebased}",
                    )
            elif expected_kind == ArtifactKind.GEO_XLSX:
                result = validate_geo_xlsx(file_path=rebased, geo_template_path=geo_template_path)
            elif expected_kind == ArtifactKind.NFCORE_RNASEQ_CSV:
                result = validate_nfcore_rnaseq_csv(file_path=rebased)
            elif expected_kind == ArtifactKind.NFCORE_SCRNASEQ_CSV:
                result = validate_nfcore_scrnaseq_csv(file_path=rebased)
            elif expected_kind == ArtifactKind.SVG_CHART:
                result = validate_svg(file_path=rebased)
            else:
                result = ValidatorResult(
                    status=ArtifactStatus.Indeterminate,
                    parser_used="none",
                    parse_success=None,
                    validation_notes=f"no validator for kind={expected_kind.value}",
                )

            # Pass-5 D2 fix: locked DD-16 / DD-24 — when the file exists but
            # `artifact_accessible == False` (i.e., permission-denied for read),
            # the row MUST report `Inaccessible`, NOT whatever the kind-specific
            # validator returned (which is often `Unreadable` from openpyxl's
            # `PermissionError` or `validate_svg`'s `(ET.ParseError, OSError)`
            # catch). `Inaccessible` is a distinct enum value precisely so Stage B
            # can rank "file present but unreadable due to permissions" worse
            # than "file present but corrupt" (DD-24 ordering: `Inaccessible >
            # Unreadable`). This override applies only when the file exists;
            # `artifact_accessible` is forced to False above when the file is
            # absent, so the guard is symmetric.
            if artifact_exists and not artifact_accessible:
                result = ValidatorResult(
                    status=ArtifactStatus.Inaccessible,
                    parser_used=result.parser_used,
                    parse_success=result.parse_success,
                    validation_notes=(
                        result.validation_notes
                        or f"file present but not readable (os.access R_OK=False): {rebased}"
                    ),
                )

            # Pass-5 D1 fix: locked DD-36 — `runtime_success == false` AND a file
            # exists on disk that the validator returned as `Valid` or
            # `Incomplete` → override to `PartialAfterFailure`. Per DD-36, "if the
            # file is on disk but unreadable / schema_invalid AND runtime failed,
            # the disk-side problem wins (e.g., `Unreadable`)" — so disk-side
            # failure statuses (`Unreadable`, `SchemaInvalid`, `Inaccessible`,
            # `Indeterminate`) are NOT replaced. This override runs AFTER the D2
            # `Inaccessible` override above so that a `runtime_success==false` +
            # `artifact_accessible==False` row lands at `Inaccessible` (disk-side
            # wins), not `PartialAfterFailure`.
            if (
                artifact_exists
                and not runtime_success
                and result.status in (ArtifactStatus.Valid, ArtifactStatus.Incomplete)
            ):
                result = ValidatorResult(
                    status=ArtifactStatus.PartialAfterFailure,
                    parser_used=result.parser_used,
                    parse_success=result.parse_success,
                    sheet_count=result.sheet_count,
                    row_count=result.row_count,
                    column_count=result.column_count,
                    nonempty_cell_count=result.nonempty_cell_count,
                    null_cell_fraction=result.null_cell_fraction,
                    required_fields_present=result.required_fields_present,
                    required_fields_complete=result.required_fields_complete,
                    missing_required_fields=result.missing_required_fields,
                    all_required_rows_complete=result.all_required_rows_complete,
                    validation_notes=result.validation_notes,
                )

        artifact_success = result.status == ArtifactStatus.Valid

        rows.append([
            run_id,
            qid,
            task_family,
            eval_id,
            artifact_expected,
            expected_kind.value,
            artifact_declared,
            manifest_path_str if artifact_declared else "",
            basename if artifact_declared else "",
            (Path(basename).suffix.lower() if basename else ""),
            runtime_success,
            failure_mode,
            artifact_exists,
            artifact_accessible,
            file_size,
            result.parser_used or "",
            result.parse_success,
            result.sheet_count,
            result.row_count,
            result.column_count,
            result.nonempty_cell_count,
            result.null_cell_fraction,
            result.required_fields_present,
            result.required_fields_complete,
            result.missing_required_fields or "",
            result.all_required_rows_complete,
            result.status.value,
            artifact_success,
            result.validation_notes,
        ])

    # Bulk-rebase-failure gate (DD-29 step 5).
    bulk_failure = rebase_attempted > 0 and rebase_failed == rebase_attempted
    if bulk_failure:
        sys.stderr.write(
            f"ERROR: {rebase_failed} of {rebase_attempted} expected artifacts failed to "
            f"rebase against {artifact_root}; check that the artifact-root path is correct "
            f"and points at a directory whose immediate children are <query_id>/ subdirectories.\n"
        )

    out_csv_path.parent.mkdir(parents=True, exist_ok=True)
    with out_csv_path.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(CSV_HEADER_29)
        for row in rows:
            w.writerow(["" if v is None else v for v in row])

    if bulk_failure and not ignore_rebase_failures:
        return 1
    return 0


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="tools.hibayes.artifact_validator",
        description=(
            "Stage A — HiBayes artifact validity validator. Reads manifest.json, "
            "rebases per-summary artifact paths against --artifact-root per locked DD-29, "
            "runs per-task-family structural validators, and emits a 29-column CSV.\n\n"
            + DD25_TEMPDIR_WARNING
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--manifest-path", required=True, type=Path)
    parser.add_argument("--artifact-root", required=True, type=Path)
    parser.add_argument(
        "--geo-template-path",
        type=Path,
        default=Path("tools/hibayes/resources/GEO-updated.json"),
    )
    parser.add_argument(
        "--out-csv",
        type=Path,
        default=Path("out/hibayes_artifact_validity.csv"),
    )
    parser.add_argument(
        "--ignore-rebase-failures",
        action="store_true",
        help="Override the non-zero exit when every declared artifact fails to rebase.",
    )
    args = parser.parse_args(argv)

    # Locked DD-39: "If the file is not found at the resolved path, Stage A errors
    # with a clear message." Fail-loud on a missing GEO template path — regardless
    # of whether the user supplied --geo-template-path or accepted the default.
    # D3 fix (MED): plan-DD-02 (= locked DD-39 Option (a)) flipped the default
    # path to the in-repo location `tools/hibayes/resources/GEO-updated.json`,
    # committed by task T0.3. The remediation text MUST reference that path /
    # T0.3, NOT the stale `make sync-vendor-deps` instruction (which targeted
    # the Option (b) credential-path resolution that was REJECTED in DD-02).
    if not args.geo_template_path.is_file():
        parser.error(
            f"--geo-template-path: file not found at {args.geo_template_path}. "
            f"Expected the in-repo committed copy at "
            f"`tools/hibayes/resources/GEO-updated.json` (per plan-DD-02 / locked "
            f"DD-39 Option (a)). Ensure task T0.3 has run (it commits the in-repo "
            f"copy) or pass an existing path explicitly via --geo-template-path."
        )

    return run_stage_a(
        manifest_path=args.manifest_path,
        artifact_root=args.artifact_root,
        geo_template_path=args.geo_template_path,
        out_csv_path=args.out_csv,
        ignore_rebase_failures=args.ignore_rebase_failures,
    )


if __name__ == "__main__":
    sys.exit(main())
