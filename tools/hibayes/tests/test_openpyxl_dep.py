"""tools/hibayes/tests/test_openpyxl_dep.py — host + in-image import smoke for T0.5.

Per locked DD-19/DD-34/DD-38 + plan-DD-07 + DL-004. Verifies openpyxl is importable
on both the host venv AND inside `hibayes-runtime-reliability:dev`.
"""
from __future__ import annotations

import shutil
import subprocess

import pytest


def test_openpyxl_importable_on_host() -> None:
    """uv add openpyxl --group tools must result in importable openpyxl on host venv."""
    import openpyxl

    assert hasattr(openpyxl, "__version__")
    assert openpyxl.__version__  # non-empty


def test_openpyxl_load_workbook_callable_on_host() -> None:
    """openpyxl.load_workbook is the primary API Stage A's GEO validator uses."""
    from openpyxl import load_workbook

    assert callable(load_workbook)


@pytest.mark.skipif(shutil.which("docker") is None, reason="docker not available")
def test_openpyxl_importable_in_image() -> None:
    """Dockerfile.hibayes-eval:47-52 extension must result in importable openpyxl in-image.

    Skips when docker is not available (CI environments without docker; developer
    machines without docker installed). Runs the image's openpyxl import as a
    smoke check.
    """
    result = subprocess.run(
        [
            "docker",
            "run",
            "--rm",
            "--platform",
            "linux/amd64",
            "hibayes-runtime-reliability:dev",
            "uv",
            "run",
            "python",
            "-c",
            "import openpyxl; print(openpyxl.__version__)",
        ],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode != 0 and "Unable to find image" in result.stderr:
        pytest.skip("hibayes-runtime-reliability:dev image not built; run `make hibayes-eval-build` first")
    assert result.returncode == 0, (
        f"in-image openpyxl import failed:\nstdout={result.stdout!r}\nstderr={result.stderr!r}"
    )
    assert result.stdout.strip(), "expected openpyxl version on stdout"


def test_pyproject_declares_openpyxl_in_tools_group() -> None:
    """`uv add openpyxl --group tools` must place openpyxl in [dependency-groups] tools.

    Parses pyproject.toml with stdlib tomllib and asserts the canonical key path
    `dependency-groups.tools` contains an openpyxl entry. A bare substring check
    is insufficient — openpyxl could be misplaced into [project.dependencies] or
    a different group and still satisfy `"openpyxl" in pyproject_text`.
    """
    import tomllib
    from pathlib import Path

    repo_root = Path(__file__).resolve().parents[3]
    with (repo_root / "pyproject.toml").open("rb") as fh:
        data = tomllib.load(fh)

    tools_group = data["dependency-groups"]["tools"]
    # PEP 735 dependency groups are a list of PEP 508 requirement strings, e.g.
    # ["pandas>=3.0.2", "openpyxl>=3.1.5"]. Match the entry whose requirement
    # name (before any version specifier / marker / extras bracket) is exactly
    # "openpyxl".
    def _req_name(req: str) -> str:
        # Strip leading whitespace, then take everything up to the first char
        # that ends the project-name token: <, >, =, !, ~, ;, [, or space.
        token = req.strip()
        for i, ch in enumerate(token):
            if ch in "<>=!~;[ ":
                return token[:i]
        return token

    names = [_req_name(req) for req in tools_group if isinstance(req, str)]
    assert "openpyxl" in names, (
        f"openpyxl not declared in [dependency-groups] tools; saw: {names!r}"
    )


def test_dockerfile_declares_openpyxl_in_eval_group() -> None:
    """DL-004: Dockerfile.hibayes-eval:47-52 RUN uv add line must include openpyxl."""
    from pathlib import Path

    repo_root = Path(__file__).resolve().parents[3]
    dockerfile = (repo_root / "Dockerfile.hibayes-eval").read_text(encoding="utf-8")
    # The `RUN uv add --no-sync --group eval` line must list openpyxl.
    # Locate the RUN block and verify openpyxl appears within it.
    assert "uv add --no-sync --group eval" in dockerfile
    # Find the line containing the uv add directive, then verify openpyxl is in the
    # multi-line continuation that follows.
    eval_run_start = dockerfile.find("RUN uv add --no-sync --group eval")
    assert eval_run_start >= 0
    # The block ends at the next `RUN ` line (or EOF). Pull a generous slice.
    block_end = dockerfile.find("\nRUN ", eval_run_start + 1)
    block = dockerfile[eval_run_start: block_end if block_end > 0 else len(dockerfile)]
    assert "openpyxl" in block, (
        f"Dockerfile RUN uv add block missing openpyxl. Block:\n{block}"
    )
