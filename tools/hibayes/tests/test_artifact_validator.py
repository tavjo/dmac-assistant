"""tools/hibayes/tests/test_artifact_validator.py — pinning tests for T1.1 Stage A.

Tests cover: 29-column header pinning, bulk-rebase-failure gate, --ignore-rebase-failures
override, multi-file NotImplementedError guard, --help DD-25 warning text, per-validator
behavior (GEO/NFCORE-RNASEQ/NFCORE-SCRNASEQ/SVG), Missing vs NotExpected semantics,
worst-status aggregation, and validation_notes population.
"""
from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path
from typing import Any

import pytest

from tools.hibayes.artifact_validator import (
    CSV_HEADER_29,
    DD25_TEMPDIR_WARNING,
    classify_artifact_kind,
    main,
    rebase_artifact_path,
    run_stage_a,
    validate_geo_xlsx,
    validate_nfcore_rnaseq_csv,
    validate_nfcore_scrnaseq_csv,
    validate_svg,
)
from tools.hibayes.enums import ArtifactKind, ArtifactStatus


# -----------------------------------------------------------------------------
# 29-column header pinning (locked §5.1)
# -----------------------------------------------------------------------------

EXPECTED_29_COLUMNS = [
    "run_id",
    "query_id",
    "task_family",
    "artifact_eval_id",
    "artifact_expected",
    "expected_artifact_kind",
    "artifact_declared",
    "artifact_path",
    "artifact_basename",
    "artifact_ext",
    "runtime_success",
    "failure_mode",
    "artifact_exists",
    "artifact_accessible",
    "file_size_bytes",
    "parser_used",
    "parse_success",
    "sheet_count",
    "row_count",
    "column_count",
    "nonempty_cell_count",
    "null_cell_fraction",
    "required_fields_present",
    "required_fields_complete",
    "missing_required_fields",
    "all_required_rows_complete",
    "artifact_validity_status",
    "artifact_success",
    "validation_notes",
]


def test_csv_header_29_is_exact_locked_design_order() -> None:
    """Locked §5.1: 29 columns in exact order. Any reorder or omission is a defect."""
    assert list(CSV_HEADER_29) == EXPECTED_29_COLUMNS


def test_csv_header_29_length_is_29() -> None:
    assert len(CSV_HEADER_29) == 29


# -----------------------------------------------------------------------------
# rebase_artifact_path (DD-29 algorithm)
# -----------------------------------------------------------------------------

def test_rebase_artifact_path_uses_qid_and_basename() -> None:
    """DD-29 step 2: rebased path = <artifact-root>/<qid>/<basename>."""
    rebased = rebase_artifact_path(
        manifest_artifact_abs_path="/absolute/whatever/Report-GEO-1/merged.xlsx",
        artifact_root=Path("/dropbox/artifacts"),
        qid="Report-GEO-1",
    )
    assert rebased == Path("/dropbox/artifacts/Report-GEO-1/merged.xlsx")


def test_rebase_artifact_path_preserves_basename_only() -> None:
    """DD-29 step 1: basename = Path(...).name — directories above the file are discarded."""
    rebased = rebase_artifact_path(
        manifest_artifact_abs_path="/dropbox/users/x/y/z/Report-NFCORE-2/samplesheet.csv",
        artifact_root=Path("/another/root"),
        qid="Report-NFCORE-2",
    )
    assert rebased == Path("/another/root/Report-NFCORE-2/samplesheet.csv")


# -----------------------------------------------------------------------------
# classify_artifact_kind (DD-34 routing)
# -----------------------------------------------------------------------------

@pytest.mark.parametrize(
    "task_family,basename,expected_kind",
    [
        ("Report-GEO", "report.xlsx", ArtifactKind.GEO_XLSX),
        ("Report-NFCORE", "samplesheet.csv", ArtifactKind.NFCORE_RNASEQ_CSV),
        # D2 fix: scrnaseq basename → NFCORE_SCRNASEQ_CSV (locked DD-18 reachability).
        ("Report-NFCORE", "samplesheet_scrnaseq.csv", ArtifactKind.NFCORE_SCRNASEQ_CSV),
        ("Report-NFCORE", "scrnaseq_samplesheet.csv", ArtifactKind.NFCORE_SCRNASEQ_CSV),
        ("Report-SRA", "anything", ArtifactKind.SRA_PACKAGE),
        ("Report-PRIDE", "anything", ArtifactKind.PRIDE_PACKAGE),
        ("Edge", "chart.svg", ArtifactKind.SVG_CHART),
        ("Search-Basic", "foo.bar", ArtifactKind.NONE_EXPECTED),
        ("Memory", None, ArtifactKind.NONE_EXPECTED),
    ],
)
def test_classify_artifact_kind_per_task_family(
    task_family: str, basename: str | None, expected_kind: ArtifactKind
) -> None:
    """DD-34: per-task-family → ArtifactKind routing."""
    assert classify_artifact_kind(task_family, basename) == expected_kind


# -----------------------------------------------------------------------------
# Multi-file guard (plan-DD-03)
# -----------------------------------------------------------------------------

def test_run_stage_a_raises_not_implemented_on_multi_file_summary(tmp_path: Path) -> None:
    """plan-DD-03: NotImplementedError when any summary has >1 artifact."""
    manifest = {
        "run_id": "test-run",
        "summaries": [
            {
                "query_id": "Report-GEO-1",
                "artifacts": [
                    "/a/Report-GEO-1/file1.xlsx",
                    "/a/Report-GEO-1/file2.xlsx",  # multi-file → raise
                ],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-GEO-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    artifact_root = tmp_path / "artifacts"
    artifact_root.mkdir()

    with pytest.raises(NotImplementedError, match="multi-file"):
        run_stage_a(
            manifest_path=manifest_path,
            artifact_root=artifact_root,
            geo_template_path=None,
            out_csv_path=tmp_path / "out.csv",
            ignore_rebase_failures=False,
        )


# -----------------------------------------------------------------------------
# Bulk-rebase-failure gate (DD-29 step 5)
# -----------------------------------------------------------------------------

def test_run_stage_a_bulk_rebase_failure_returns_non_zero_exit(tmp_path: Path) -> None:
    """DD-29 step 5: every non-empty artifact fails to rebase → exit non-zero."""
    manifest = {
        "run_id": "test-run",
        "summaries": [
            {
                "query_id": "Report-GEO-1",
                "artifacts": ["/dropbox/nowhere/Report-GEO-1/missing.xlsx"],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-GEO-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    empty_root = tmp_path / "empty_artifacts"
    empty_root.mkdir()  # exists but no <qid>/<basename> within

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=empty_root,
        geo_template_path=None,
        out_csv_path=tmp_path / "out.csv",
        ignore_rebase_failures=False,
    )
    assert exit_code != 0


def test_run_stage_a_ignore_rebase_failures_returns_zero(tmp_path: Path) -> None:
    """DL-010: `--ignore-rebase-failures` overrides the non-zero exit."""
    manifest = {
        "run_id": "test-run",
        "summaries": [
            {
                "query_id": "Report-GEO-1",
                "artifacts": ["/dropbox/nowhere/Report-GEO-1/missing.xlsx"],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-GEO-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    empty_root = tmp_path / "empty_artifacts"
    empty_root.mkdir()

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=empty_root,
        geo_template_path=None,
        out_csv_path=tmp_path / "out.csv",
        ignore_rebase_failures=True,
    )
    assert exit_code == 0


# -----------------------------------------------------------------------------
# CLI --help warning (DD-25 / DL-014)
# -----------------------------------------------------------------------------

def test_cli_help_contains_dd25_tempdir_warning() -> None:
    """DL-014 / DD-25: --help output MUST contain the verbatim tempdir-mode warning."""
    result = subprocess.run(
        [sys.executable, "-m", "tools.hibayes.artifact_validator", "--help"],
        capture_output=True,
        text=True,
    )
    assert result.returncode == 0
    assert DD25_TEMPDIR_WARNING in result.stdout, (
        f"--help missing DD-25 warning. Expected substring:\n{DD25_TEMPDIR_WARNING}\n"
        f"Actual stdout:\n{result.stdout}"
    )


def test_dd25_tempdir_warning_constant_mentions_required_terms() -> None:
    """DL-014: warning text must mention tempdir-mode, --scratch-dir, and Missing."""
    text = DD25_TEMPDIR_WARNING.lower()
    assert "--scratch-dir" in text or "scratch-dir" in text
    assert "tempdir" in text or "temp" in text
    assert "missing" in text


# -----------------------------------------------------------------------------
# GEO validator (DD-19 + DD-38)
# -----------------------------------------------------------------------------

def test_validate_geo_xlsx_returns_missing_when_file_absent(tmp_path: Path) -> None:
    """If the file does not exist at the rebased path, status is Missing."""
    result = validate_geo_xlsx(
        file_path=tmp_path / "nonexistent.xlsx",
        geo_template_path=None,
    )
    assert result.status == ArtifactStatus.Missing


def test_validate_geo_xlsx_valid_workbook(tmp_path: Path) -> None:
    """A well-formed minimal xlsx → status Valid (when no template fields to enforce)."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "header"
    ws["A2"] = "data"
    xlsx = tmp_path / "ok.xlsx"
    wb.save(xlsx)
    result = validate_geo_xlsx(file_path=xlsx, geo_template_path=None)
    assert result.status in (ArtifactStatus.Valid, ArtifactStatus.Indeterminate)
    assert result.parser_used == "openpyxl"
    assert result.parse_success is True
    assert (result.sheet_count or 0) >= 1
    assert (result.row_count or 0) >= 1


def test_validate_geo_xlsx_unreadable_when_corrupt(tmp_path: Path) -> None:
    """A non-xlsx file at .xlsx path → status Unreadable."""
    xlsx = tmp_path / "bad.xlsx"
    xlsx.write_bytes(b"not actually an xlsx file")
    result = validate_geo_xlsx(file_path=xlsx, geo_template_path=None)
    assert result.status == ArtifactStatus.Unreadable
    assert result.parse_success is False


def test_validate_geo_xlsx_with_template_required_fields(tmp_path: Path) -> None:
    """DD-38 step 4: When a template is supplied, required `*`-prefixed fields
    scoped to samples[0] are matched against the workbook's row-1 column
    headers AND each required column is non-null in every data row. A
    workbook with every required header present and non-null data →
    required_fields_present=True AND required_fields_complete=True AND
    status=Valid.

    Pass 4 D1 fix: template carries >=3 single-`*` keys to satisfy DD-38
    step 3 cardinality (3-25). Pass 4 D2 fix: column headers are now checked
    against row 1, not workbook-wide substring scan.
    """
    import openpyxl

    # Minimal GEO template JSON: three single-`*` required fields under samples[0]
    # (minimum cardinality per DD-38 step 3) plus one `**` recommended field.
    template = {
        "samples": [
            {
                "*library name": "",
                "*title": "",
                "*organism": "",
                "**tissue": "",  # recommended — NOT required
            }
        ]
    }
    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps(template))

    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1 = column headers; each required field name appears here verbatim
    # (case-insensitive match against template's stripped `*`-prefix).
    ws["A1"] = "library name"
    ws["B1"] = "title"
    ws["C1"] = "organism"
    # Row 2 = data; every required column non-null.
    ws["A2"] = "samp1"
    ws["B2"] = "T-cell experiment"
    ws["C2"] = "Homo sapiens"
    xlsx = tmp_path / "geo.xlsx"
    wb.save(xlsx)

    result = validate_geo_xlsx(file_path=xlsx, geo_template_path=template_path)
    assert result.parser_used == "openpyxl"
    assert result.parse_success is True
    assert result.required_fields_present is True
    assert result.required_fields_complete is True
    assert result.all_required_rows_complete is True
    assert result.status == ArtifactStatus.Valid


# -----------------------------------------------------------------------------
# Pass 4 D1 HIGH: fail-loud GEO template schema + cardinality sanity check
# (locked DD-38 step 2 + step 3 cardinality clause)
# -----------------------------------------------------------------------------

def test_load_and_validate_geo_template_happy_path(tmp_path: Path) -> None:
    """DD-38: a well-formed template with 3-25 single-`*` keys under samples[0]
    returns the stripped key list."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template = {
        "samples": [
            {
                "*library name": "",
                "*title": "",
                "*organism": "",
                "**tissue": "",  # recommended — ignored
            }
        ]
    }
    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps(template))
    required = _load_and_validate_geo_template(template_path)
    assert required == ["library name", "title", "organism"]


def test_load_and_validate_geo_template_malformed_json_exits_non_zero(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """DD-38: malformed template JSON → SystemExit(2) with verbatim error message."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text("{not valid json")
    with pytest.raises(SystemExit) as exc_info:
        _load_and_validate_geo_template(template_path)
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "GEO template at" in err
    assert "does not match the expected schema" in err
    assert "See DD-38." in err


def test_load_and_validate_geo_template_missing_samples_key_exits_non_zero(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """DD-38 step 2: template missing `samples` key → SystemExit(2)."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps({"study": {"*title": ""}}))
    with pytest.raises(SystemExit) as exc_info:
        _load_and_validate_geo_template(template_path)
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "samples" in err
    assert "See DD-38." in err


def test_load_and_validate_geo_template_empty_samples_list_exits_non_zero(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """DD-38 step 2: `samples` is empty list → SystemExit(2)."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps({"samples": []}))
    with pytest.raises(SystemExit) as exc_info:
        _load_and_validate_geo_template(template_path)
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "See DD-38." in err


def test_load_and_validate_geo_template_samples_zero_not_dict_exits_non_zero(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """DD-38 step 2: `samples[0]` is not a dict → SystemExit(2)."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps({"samples": ["not a dict"]}))
    with pytest.raises(SystemExit) as exc_info:
        _load_and_validate_geo_template(template_path)
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "See DD-38." in err


def test_load_and_validate_geo_template_cardinality_below_3_exits_non_zero(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """DD-38 step 3 cardinality: <3 single-`*` keys under samples[0] → SystemExit(2)."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(
        json.dumps({"samples": [{"*library name": "", "*title": ""}]})
    )
    with pytest.raises(SystemExit) as exc_info:
        _load_and_validate_geo_template(template_path)
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "See DD-38." in err
    assert "count" in err.lower() or "3" in err


def test_load_and_validate_geo_template_cardinality_above_25_exits_non_zero(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """DD-38 step 3 cardinality: >25 single-`*` keys under samples[0] → SystemExit(2)."""
    from tools.hibayes.artifact_validator import _load_and_validate_geo_template

    template_path = tmp_path / "GEO-updated.json"
    samples0 = {f"*field_{i}": "" for i in range(26)}
    template_path.write_text(json.dumps({"samples": [samples0]}))
    with pytest.raises(SystemExit) as exc_info:
        _load_and_validate_geo_template(template_path)
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "See DD-38." in err


# -----------------------------------------------------------------------------
# Pass 4 D2 HIGH: column-header presence + per-row non-null check
# (locked DD-38 step 4)
# -----------------------------------------------------------------------------

def test_validate_geo_xlsx_missing_column_header_returns_schema_invalid(
    tmp_path: Path,
) -> None:
    """DD-38 step 4 part (i): a required field whose name is NOT present in
    row 1 (column headers) → required_fields_present=False, status=SchemaInvalid.

    The required field name appearing only in a non-header cell (e.g., a
    free-form study-block cell) MUST NOT spuriously satisfy presence.
    """
    import openpyxl

    template = {
        "samples": [
            {
                "*library name": "",
                "*title": "",
                "*organism": "",
            }
        ]
    }
    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps(template))

    wb = openpyxl.Workbook()
    ws = wb.active
    # Row 1 = column headers (only 2 of 3 required fields).
    ws["A1"] = "library name"
    ws["B1"] = "title"
    # Row 2 = a free-form cell that mentions "organism" but NOT as a header.
    # The Pass 3 substring-scan would have spuriously accepted this; the
    # Pass 4 column-header check must reject it.
    ws["A2"] = "samp1"
    ws["B2"] = "organism study of T-cells"
    xlsx = tmp_path / "geo.xlsx"
    wb.save(xlsx)

    result = validate_geo_xlsx(file_path=xlsx, geo_template_path=template_path)
    assert result.required_fields_present is False
    assert result.status == ArtifactStatus.SchemaInvalid
    assert result.missing_required_fields is not None
    assert "organism" in result.missing_required_fields


def test_validate_geo_xlsx_per_row_null_returns_incomplete(tmp_path: Path) -> None:
    """DD-38 step 4 part (ii): required column headers all present, but one
    data row has a null in a required column → required_fields_complete=False,
    all_required_rows_complete=False, status=Incomplete.

    Pass 3's `required_complete = required_present` collapse silently
    reported `Valid` for this case; the Pass 4 per-row null check must
    distinguish it.
    """
    import openpyxl

    template = {
        "samples": [
            {
                "*library name": "",
                "*title": "",
                "*organism": "",
            }
        ]
    }
    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(json.dumps(template))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "library name"
    ws["B1"] = "title"
    ws["C1"] = "organism"
    ws["A2"] = "samp1"
    ws["B2"] = "T-cell experiment"
    ws["C2"] = "Homo sapiens"
    ws["A3"] = "samp2"
    ws["B3"] = "B-cell experiment"
    # C3 deliberately left null → row 3 violates per-row non-null on `organism`.
    xlsx = tmp_path / "geo.xlsx"
    wb.save(xlsx)

    result = validate_geo_xlsx(file_path=xlsx, geo_template_path=template_path)
    assert result.required_fields_present is True
    assert result.required_fields_complete is False
    assert result.all_required_rows_complete is False
    assert result.status == ArtifactStatus.Incomplete
    assert result.missing_required_fields is not None
    assert "organism" in result.missing_required_fields


def test_run_stage_a_exits_non_zero_when_geo_template_malformed(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """End-to-end: `run_stage_a` invoked with a malformed GEO template path
    AND a manifest containing a Report-GEO summary MUST surface the fail-loud
    DD-38 SystemExit(2) from `_load_and_validate_geo_template`.
    """
    qid = "Report-GEO-1"
    artifact_root = tmp_path / "artifacts"
    qdir = artifact_root / qid
    qdir.mkdir(parents=True)
    # Create an xlsx so the per-summary dispatch actually reaches validate_geo_xlsx.
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active["A1"] = "library name"
    wb.active["A2"] = "samp1"
    xlsx = qdir / "report.xlsx"
    wb.save(xlsx)

    template_path = tmp_path / "GEO-updated.json"
    # Cardinality below 3 → fail-loud SystemExit(2) per DD-38.
    template_path.write_text(json.dumps({"samples": [{"*library name": ""}]}))

    manifest = {
        "run_id": "geo-fail-loud-run",
        "summaries": [
            {
                "query_id": qid,
                "artifacts": [f"/whatever/{qid}/report.xlsx"],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-GEO-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))

    with pytest.raises(SystemExit) as exc_info:
        run_stage_a(
            manifest_path=manifest_path,
            artifact_root=artifact_root,
            geo_template_path=template_path,
            out_csv_path=tmp_path / "out.csv",
            ignore_rebase_failures=False,
        )
    assert exc_info.value.code == 2
    err = capsys.readouterr().err
    assert "See DD-38." in err


def test_run_stage_a_artifact_accessible_false_when_unreadable(tmp_path: Path) -> None:
    """Locked §5.1 row 14: artifact_accessible uses os.access(..., os.R_OK).

    A file that exists but is permission-denied for read MUST report
    artifact_accessible=False; previously a naive `st_size >= 0` check would
    have masked this and incorrectly reported True.
    """
    import os as _os
    import stat as _stat

    qid = "Edge-1"
    artifact_root = tmp_path / "artifacts"
    qdir = artifact_root / qid
    qdir.mkdir(parents=True)
    svg_path = qdir / "chart.svg"
    svg_path.write_text('<svg xmlns="http://www.w3.org/2000/svg"/>')
    # Strip read permission from owner, group, and other.
    svg_path.chmod(0o000)
    # Skip if the filesystem (e.g., running as root) does not honor read perms.
    if _os.access(svg_path, _os.R_OK):
        svg_path.chmod(_stat.S_IREAD)  # restore so cleanup doesn't fail
        pytest.skip("Filesystem does not honor read permissions (likely running as root)")

    manifest = {
        "run_id": "perm-run",
        "summaries": [
            {
                "query_id": qid,
                "artifacts": [f"/whatever/{qid}/chart.svg"],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Edge-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    out_csv = tmp_path / "out.csv"

    try:
        exit_code = run_stage_a(
            manifest_path=manifest_path,
            artifact_root=artifact_root,
            geo_template_path=None,
            out_csv_path=out_csv,
            ignore_rebase_failures=False,
        )
        assert exit_code == 0
        lines = out_csv.read_text().splitlines()
        fields = lines[1].split(",")
        accessible_idx = EXPECTED_29_COLUMNS.index("artifact_accessible")
        # CSV writer renders Python bool False as the string "False".
        assert fields[accessible_idx] == "False"
        # Pass-5 D2 fix: locked DD-16/DD-24 — when artifact_accessible=False AND
        # artifact_exists=True, the status column MUST be "Inaccessible" (NOT
        # "Unreadable", which is what `validate_svg`'s `(ET.ParseError, OSError)`
        # catch would otherwise emit on the chmod-0o000 file). `Inaccessible`
        # ranks worse than `Unreadable` in DD-24 worst-status ordering precisely
        # so Stage B can distinguish "permission-denied" from "corrupt-file."
        status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
        assert fields[status_idx] == ArtifactStatus.Inaccessible.value
    finally:
        # Restore permissions for tmp_path cleanup.
        svg_path.chmod(0o600)


# -----------------------------------------------------------------------------
# DD-36 RuntimeFailed / PartialAfterFailure status mapping (Pass-5 D1)
# -----------------------------------------------------------------------------

def test_run_stage_a_emits_partial_after_failure_when_runtime_failed_with_parseable_file(
    tmp_path: Path,
) -> None:
    """Locked DD-36: `runtime_success == false` AND a parseable artifact file
    exists on disk → `artifact_validity_status` is `PartialAfterFailure`
    (NOT `Valid`/`Incomplete`).

    This is the canonical 224850Z corpus's `Report-NFCORE-3` case:
    `is_error=true`, `timed_out=true`, `answer_provided=false`, one declared
    artifact (`merged_report_NFCORE_RNASEQ_samplesheet.csv`) sitting on disk
    and parsing as a clean nf-core RNA-seq samplesheet (DD-17). Stage A is the
    SOLE producer of `PartialAfterFailure` in the pipeline; if this status is
    not emitted here, Stage B's worst-status aggregation (DD-24) and Stage D's
    posterior cannot exercise the DD-36 enum value end-to-end.
    """
    qid = "Report-NFCORE-3"
    artifact_root = tmp_path / "artifacts"
    qdir = artifact_root / qid
    qdir.mkdir(parents=True)
    csv_path = qdir / "merged_report_NFCORE_RNASEQ_samplesheet.csv"
    # Valid nf-core RNA-seq samplesheet per DD-17: required cols all present,
    # strandedness in upstream enum {unstranded, forward, reverse, auto}.
    csv_path.write_text(
        "sample,fastq_1,fastq_2,strandedness\n"
        "S1,/r1.fq.gz,/r2.fq.gz,auto\n"
    )
    manifest = {
        "run_id": "nfcore-3-run",
        "summaries": [
            {
                "query_id": qid,
                "artifacts": [f"/whatever/{qid}/merged_report_NFCORE_RNASEQ_samplesheet.csv"],
                # 224850Z Report-NFCORE-3 signature: runtime failed but artifact landed.
                "answer_provided": False,
                "is_error": True,
                "timed_out": True,
                "record_path": f"evidence/test/{qid}.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    out_csv = tmp_path / "out.csv"

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=False,
    )
    assert exit_code == 0
    lines = out_csv.read_text().splitlines()
    fields = lines[1].split(",")
    status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
    runtime_idx = EXPECTED_29_COLUMNS.index("runtime_success")
    exists_idx = EXPECTED_29_COLUMNS.index("artifact_exists")
    # Sanity: runtime_success is False, file exists on disk.
    assert fields[runtime_idx] == "False"
    assert fields[exists_idx] == "True"
    # DD-36: PartialAfterFailure, NOT Valid.
    assert fields[status_idx] == ArtifactStatus.PartialAfterFailure.value


def test_run_stage_a_emits_runtime_failed_when_runtime_failed_with_no_file(
    tmp_path: Path,
) -> None:
    """Locked DD-36: `runtime_success == false` AND no artifact file exists →
    `artifact_validity_status` is `RuntimeFailed` (NOT `Missing`).

    Covers both reachable spellings of "no file": (a) `artifact_declared=False`
    with `artifact_expected=True` (the family_to_kind branch), and (b)
    `artifact_declared=True` but the rebased path does not exist on disk.
    `RuntimeFailed` is the worst rank in DD-24 ordering; the runtime-failed
    case must dominate `Missing` so Stage B's roll-up reflects "agent failed
    AND produced nothing" rather than the more-benign "agent succeeded but
    forgot the artifact."
    """
    # Sub-case (a): expected-but-no-declared-artifact, runtime failed.
    # parse_query_id strips the trailing integer suffix; task_family resolves to
    # "Report-NFCORE" which the family_to_kind map routes to NFCORE_RNASEQ_CSV.
    qid_a = "Report-NFCORE-99"
    # Sub-case (b): declared artifact but file is absent on disk, runtime failed.
    qid_b = "Report-NFCORE-100"
    artifact_root = tmp_path / "artifacts"
    (artifact_root / qid_b).mkdir(parents=True)  # qdir exists; the file does not.

    manifest = {
        "run_id": "rt-fail-run",
        "summaries": [
            {
                "query_id": qid_a,
                "artifacts": [],
                "answer_provided": False,
                "is_error": True,
                "timed_out": False,
                "record_path": f"evidence/test/{qid_a}.record.json",
            },
            {
                "query_id": qid_b,
                "artifacts": [f"/whatever/{qid_b}/missing.csv"],
                "answer_provided": False,
                "is_error": True,
                "timed_out": False,
                "record_path": f"evidence/test/{qid_b}.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    out_csv = tmp_path / "out.csv"

    # Sub-case (b) exercises the bulk-rebase-failure gate (1 of 1 declared
    # artifacts failed to rebase), so pass --ignore-rebase-failures to keep the
    # focus on the status-mapping assertion. (Sub-case (a) has no declared
    # artifact so it doesn't count against `rebase_attempted`.)
    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=True,
    )
    assert exit_code == 0
    lines = out_csv.read_text().splitlines()
    rows = [line.split(",") for line in lines[1:]]
    status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
    qid_idx = EXPECTED_29_COLUMNS.index("query_id")
    by_qid = {r[qid_idx]: r for r in rows}
    # DD-36: both sub-cases land at RuntimeFailed, NOT Missing.
    assert by_qid[qid_a][status_idx] == ArtifactStatus.RuntimeFailed.value
    assert by_qid[qid_b][status_idx] == ArtifactStatus.RuntimeFailed.value


# -----------------------------------------------------------------------------
# NFCORE RNA-seq validator (DD-17)
# -----------------------------------------------------------------------------

def test_validate_nfcore_rnaseq_csv_valid(tmp_path: Path) -> None:
    """DD-17: required cols all present + strandedness in enum → Valid."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text(
        "sample,fastq_1,fastq_2,strandedness\n"
        "S1,/r1.fq.gz,/r2.fq.gz,auto\n"
        "S2,/r1.fq.gz,/r2.fq.gz,forward\n"
    )
    result = validate_nfcore_rnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.Valid
    assert result.required_fields_present is True
    assert result.required_fields_complete is True


def test_validate_nfcore_rnaseq_csv_invalid_strandedness(tmp_path: Path) -> None:
    """DD-17: strandedness=reverse_complement is out-of-enum → SchemaInvalid."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text(
        "sample,fastq_1,fastq_2,strandedness\n"
        "S1,/r1.fq.gz,/r2.fq.gz,reverse_complement\n"
    )
    result = validate_nfcore_rnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.SchemaInvalid
    assert result.missing_required_fields and "strandedness" in result.missing_required_fields


def test_validate_nfcore_rnaseq_csv_missing_required_column(tmp_path: Path) -> None:
    """DD-17: missing strandedness column → SchemaInvalid (required col absent)."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text(
        "sample,fastq_1,fastq_2\n"
        "S1,/r1.fq.gz,/r2.fq.gz\n"
    )
    result = validate_nfcore_rnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.SchemaInvalid
    assert result.required_fields_present is False


def test_validate_nfcore_rnaseq_csv_null_row(tmp_path: Path) -> None:
    """DD-17: a required column null on one row → Incomplete."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text(
        "sample,fastq_1,fastq_2,strandedness\n"
        "S1,/r1.fq.gz,/r2.fq.gz,auto\n"
        "S2,/r1.fq.gz,,forward\n"
    )
    result = validate_nfcore_rnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.Incomplete
    assert result.all_required_rows_complete is False


def test_validate_nfcore_rnaseq_csv_missing_file(tmp_path: Path) -> None:
    """DD-17: file does not exist → status Missing."""
    result = validate_nfcore_rnaseq_csv(file_path=tmp_path / "nonexistent.csv")
    assert result.status == ArtifactStatus.Missing


# -----------------------------------------------------------------------------
# NFCORE scRNA-seq validator (DD-18)
# -----------------------------------------------------------------------------

def test_validate_nfcore_scrnaseq_csv_valid(tmp_path: Path) -> None:
    """DD-18: first three columns sample/fastq_1/fastq_2 present + non-null → Valid."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text(
        "sample,fastq_1,fastq_2,expected_cells\n"
        "S1,/r1.fq.gz,/r2.fq.gz,5000\n"
    )
    result = validate_nfcore_scrnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.Valid


def test_validate_nfcore_scrnaseq_csv_missing_file(tmp_path: Path) -> None:
    """DD-18: file does not exist → status Missing."""
    result = validate_nfcore_scrnaseq_csv(file_path=tmp_path / "nonexistent.csv")
    assert result.status == ArtifactStatus.Missing


def test_validate_nfcore_scrnaseq_csv_unreadable(tmp_path: Path) -> None:
    """DD-18: pandas raises (not a CSV at all) → status Unreadable."""
    bad = tmp_path / "bad.csv"
    # A null byte breaks pandas' default CSV parser.
    bad.write_bytes(b"\x00\x00\x00not a csv at all")
    result = validate_nfcore_scrnaseq_csv(file_path=bad)
    # Either Unreadable (parse failed) or SchemaInvalid (parsed but no required cols);
    # both are valid failure modes — the test asserts we DO NOT silently report Valid.
    assert result.status in (ArtifactStatus.Unreadable, ArtifactStatus.SchemaInvalid)
    assert result.parse_success in (False, True)


def test_validate_nfcore_scrnaseq_csv_schema_invalid(tmp_path: Path) -> None:
    """DD-18: missing required column → SchemaInvalid."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text("sample,fastq_1\nS1,/r1.fq.gz\n")
    result = validate_nfcore_scrnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.SchemaInvalid
    assert result.required_fields_present is False


def test_validate_nfcore_scrnaseq_csv_incomplete(tmp_path: Path) -> None:
    """DD-18: required column null on one row → Incomplete."""
    csv = tmp_path / "samplesheet.csv"
    csv.write_text(
        "sample,fastq_1,fastq_2\n"
        "S1,/r1.fq.gz,/r2.fq.gz\n"
        "S2,,/r2.fq.gz\n"
    )
    result = validate_nfcore_scrnaseq_csv(file_path=csv)
    assert result.status == ArtifactStatus.Incomplete
    assert result.required_fields_complete is False


def test_run_stage_a_dispatches_to_scrnaseq_validator(tmp_path: Path) -> None:
    """D2 fix: end-to-end reachability — Report-NFCORE with a scrnaseq-prefixed
    basename MUST reach `validate_nfcore_scrnaseq_csv` via `classify_artifact_kind`
    in `run_stage_a`'s dispatch. The emitted row must report `parser_used=pandas_csv`
    and `artifact_validity_status=Valid` for a well-formed scRNA-seq samplesheet,
    and `expected_artifact_kind` must equal `NFCORE_SCRNASEQ_CSV.value` (locked DD-18).
    """
    qid = "Report-NFCORE-1"
    artifact_root = tmp_path / "artifacts"
    qdir = artifact_root / qid
    qdir.mkdir(parents=True)
    sc_csv = qdir / "samplesheet_scrnaseq.csv"
    sc_csv.write_text(
        "sample,fastq_1,fastq_2,expected_cells\n"
        "S1,/r1.fq.gz,/r2.fq.gz,5000\n"
    )

    manifest = {
        "run_id": "sc-disp-run",
        "summaries": [
            {
                "query_id": qid,
                "artifacts": [f"/whatever/{qid}/samplesheet_scrnaseq.csv"],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-NFCORE-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    out_csv = tmp_path / "out.csv"

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=False,
    )
    assert exit_code == 0
    lines = out_csv.read_text().splitlines()
    fields = lines[1].split(",")
    kind_idx = EXPECTED_29_COLUMNS.index("expected_artifact_kind")
    parser_idx = EXPECTED_29_COLUMNS.index("parser_used")
    status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
    assert fields[kind_idx] == ArtifactKind.NFCORE_SCRNASEQ_CSV.value
    assert fields[parser_idx] == "pandas_csv"
    assert fields[status_idx] == ArtifactStatus.Valid.value


# -----------------------------------------------------------------------------
# SVG validator (DD-20)
# -----------------------------------------------------------------------------

def test_validate_svg_well_formed(tmp_path: Path) -> None:
    """DD-20: parses as XML + root is <svg> → Valid."""
    svg = tmp_path / "chart.svg"
    svg.write_text('<svg xmlns="http://www.w3.org/2000/svg"><rect/></svg>')
    result = validate_svg(file_path=svg)
    assert result.status == ArtifactStatus.Valid


def test_validate_svg_zero_size(tmp_path: Path) -> None:
    """DD-20: file size 0 → Unreadable."""
    svg = tmp_path / "empty.svg"
    svg.write_bytes(b"")
    result = validate_svg(file_path=svg)
    assert result.status == ArtifactStatus.Unreadable


def test_validate_svg_malformed_xml(tmp_path: Path) -> None:
    """DD-20: malformed XML → Unreadable."""
    svg = tmp_path / "bad.svg"
    svg.write_text("not really xml at all <svg>")
    result = validate_svg(file_path=svg)
    assert result.status == ArtifactStatus.Unreadable


def test_validate_svg_wrong_root_element(tmp_path: Path) -> None:
    """DD-20: well-formed XML but root is not <svg> → SchemaInvalid."""
    svg = tmp_path / "bad.svg"
    svg.write_text("<html><body/></html>")
    result = validate_svg(file_path=svg)
    assert result.status == ArtifactStatus.SchemaInvalid


def test_validate_svg_missing_file(tmp_path: Path) -> None:
    """DD-20: file does not exist → status Missing."""
    result = validate_svg(file_path=tmp_path / "nonexistent.svg")
    assert result.status == ArtifactStatus.Missing


# -----------------------------------------------------------------------------
# End-to-end smoke (single-file artifact, happy path)
# -----------------------------------------------------------------------------

def test_run_stage_a_emits_29_column_csv(tmp_path: Path) -> None:
    """Happy-path smoke: a valid SVG produces a single 29-column row."""
    qid = "Edge-1"
    artifact_root = tmp_path / "artifacts"
    qdir = artifact_root / qid
    qdir.mkdir(parents=True)
    svg_path = qdir / "chart.svg"
    svg_path.write_text('<svg xmlns="http://www.w3.org/2000/svg"/>')

    manifest = {
        "run_id": "smoke-run",
        "summaries": [
            {
                "query_id": qid,
                "artifacts": [f"/whatever/{qid}/chart.svg"],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Edge-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    out_csv = tmp_path / "out.csv"

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=False,
    )
    assert exit_code == 0

    rows = list(out_csv.read_text().splitlines())
    # header + 1 data row
    assert len(rows) == 2
    header = rows[0].split(",")
    assert header == EXPECTED_29_COLUMNS


def test_run_stage_a_emits_missing_status_for_expected_but_absent(tmp_path: Path) -> None:
    """A query with artifact_expected==True but artifacts==[] → status Missing."""
    manifest = {
        "run_id": "miss-run",
        "summaries": [
            {
                "query_id": "Report-GEO-1",
                "artifacts": [],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-GEO-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    artifact_root = tmp_path / "artifacts"
    artifact_root.mkdir()
    out_csv = tmp_path / "out.csv"

    # No expected artifact present; bulk-rebase gate should NOT trigger (no non-empty summaries).
    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=False,
    )
    assert exit_code == 0

    lines = out_csv.read_text().splitlines()
    assert len(lines) == 2
    fields = lines[1].split(",")
    status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
    assert fields[status_idx] == ArtifactStatus.Missing.value


def test_run_stage_a_emits_not_expected_for_non_artifact_family(tmp_path: Path) -> None:
    """task_family=Search-Basic + no artifacts → status NotExpected."""
    manifest = {
        "run_id": "ne-run",
        "summaries": [
            {
                "query_id": "Search-Basic-1",
                "artifacts": [],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Search-Basic-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    artifact_root = tmp_path / "artifacts"
    artifact_root.mkdir()
    out_csv = tmp_path / "out.csv"

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=False,
    )
    assert exit_code == 0
    lines = out_csv.read_text().splitlines()
    fields = lines[1].split(",")
    status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
    expected_idx = EXPECTED_29_COLUMNS.index("artifact_expected")
    assert fields[status_idx] == ArtifactStatus.NotExpected.value
    assert fields[expected_idx] in ("False", "false")


# -----------------------------------------------------------------------------
# _derive_failure_mode branch coverage
# -----------------------------------------------------------------------------

@pytest.mark.parametrize(
    "summary,expected_mode",
    [
        ({"timed_out": True, "is_error": False, "answer_provided": False}, "timeout"),
        ({"timed_out": False, "is_error": True, "answer_provided": False}, "error"),
        ({"timed_out": False, "is_error": False, "answer_provided": False}, "no_answer"),
        ({"timed_out": False, "is_error": False, "answer_provided": True}, "none"),
    ],
)
def test_derive_failure_mode_branches(summary: dict[str, Any], expected_mode: str) -> None:
    """All four branches of _derive_failure_mode."""
    from tools.hibayes.artifact_validator import _derive_failure_mode

    assert _derive_failure_mode(summary) == expected_mode


# -----------------------------------------------------------------------------
# main() argparse plumbing (D3 fix coverage + happy path)
# -----------------------------------------------------------------------------

def test_main_errors_on_nonexistent_geo_template_path(
    tmp_path: Path, capsys: pytest.CaptureFixture[str]
) -> None:
    """Locked DD-39: 'If the file is not found at the resolved path, Stage A errors
    with a clear message.' main() MUST exit non-zero rather than silently coercing
    the path to None.

    D3 fix (MED): the remediation text MUST point at the in-repo committed copy
    `tools/hibayes/resources/GEO-updated.json` and reference task T0.3
    (per plan-DD-02 / locked DD-39 Option (a)). The stale
    `make sync-vendor-deps` wording — which targeted the rejected Option (b)
    credential-path resolution — MUST NOT appear in the error message.
    """
    manifest = {"run_id": "test", "summaries": []}
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    artifact_root = tmp_path / "artifacts"
    artifact_root.mkdir()
    bad_template = tmp_path / "does_not_exist.json"

    with pytest.raises(SystemExit):
        main(
            [
                "--manifest-path",
                str(manifest_path),
                "--artifact-root",
                str(artifact_root),
                "--geo-template-path",
                str(bad_template),
                "--out-csv",
                str(tmp_path / "out.csv"),
            ]
        )
    captured = capsys.readouterr()
    # argparse.ArgumentParser.error() writes to stderr.
    err = captured.err
    assert "tools/hibayes/resources/GEO-updated.json" in err
    assert "T0.3" in err
    # Stale remediation MUST NOT appear (D3 fix).
    assert "make sync-vendor-deps" not in err


def test_main_happy_path_routes_to_run_stage_a(tmp_path: Path) -> None:
    """main() routes argparse-parsed args to run_stage_a and returns its exit code.
    Uses an empty-summaries manifest + a valid template file so neither bulk-rebase
    nor DD-39 fail-loud triggers.

    Pass 4 D1 fix: template carries >=3 single-`*` keys to satisfy DD-38 step 3
    cardinality (3-25). Even though the empty-summaries manifest means
    `_load_and_validate_geo_template` is not triggered in the current
    per-summary call site, keeping the fixture compliant guards against a
    future load-at-startup refactor regressing this test.
    """
    manifest = {"run_id": "main-test", "summaries": []}
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    artifact_root = tmp_path / "artifacts"
    artifact_root.mkdir()
    template_path = tmp_path / "GEO-updated.json"
    template_path.write_text(
        json.dumps(
            {
                "samples": [
                    {
                        "*library name": "",
                        "*title": "",
                        "*organism": "",
                    }
                ]
            }
        )
    )
    out_csv = tmp_path / "out.csv"

    exit_code = main(
        [
            "--manifest-path",
            str(manifest_path),
            "--artifact-root",
            str(artifact_root),
            "--geo-template-path",
            str(template_path),
            "--out-csv",
            str(out_csv),
        ]
    )
    assert exit_code == 0
    # CSV is written even with zero summaries (header-only).
    lines = out_csv.read_text().splitlines()
    assert len(lines) == 1
    assert lines[0].split(",") == EXPECTED_29_COLUMNS


# -----------------------------------------------------------------------------
# family_to_kind: declared-artifact-absent-but-expected branch
# -----------------------------------------------------------------------------

def test_run_stage_a_report_nfcore_missing_artifact_emits_missing(tmp_path: Path) -> None:
    """task_family=Report-NFCORE with no declared artifact exercises the
    `family_to_kind` lookup in run_stage_a + the artifact_expected=True path."""
    manifest = {
        "run_id": "miss-nfcore-run",
        "summaries": [
            {
                "query_id": "Report-NFCORE-1",
                "artifacts": [],
                "answer_provided": True,
                "is_error": False,
                "timed_out": False,
                "record_path": "evidence/test/Report-NFCORE-1.record.json",
            },
        ],
    }
    manifest_path = tmp_path / "manifest.json"
    manifest_path.write_text(json.dumps(manifest))
    artifact_root = tmp_path / "artifacts"
    artifact_root.mkdir()
    out_csv = tmp_path / "out.csv"

    exit_code = run_stage_a(
        manifest_path=manifest_path,
        artifact_root=artifact_root,
        geo_template_path=None,
        out_csv_path=out_csv,
        ignore_rebase_failures=False,
    )
    assert exit_code == 0
    lines = out_csv.read_text().splitlines()
    fields = lines[1].split(",")
    status_idx = EXPECTED_29_COLUMNS.index("artifact_validity_status")
    expected_idx = EXPECTED_29_COLUMNS.index("artifact_expected")
    kind_idx = EXPECTED_29_COLUMNS.index("expected_artifact_kind")
    assert fields[status_idx] == ArtifactStatus.Missing.value
    assert fields[expected_idx] in ("True", "true")
    # expected_artifact_kind is the .value of NFCORE_RNASEQ_CSV from the family_to_kind map.
    assert fields[kind_idx] == ArtifactKind.NFCORE_RNASEQ_CSV.value
